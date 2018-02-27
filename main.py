from openpyxl import load_workbook

class global_history:
    l_1 = {}
    l_2 = {}
    l_3 = {}
    l_4 = {}
    l_5 = {}
    def predict(self, c_history, current_virtual_move):
        if c_history.append(current_virtual_move) in l_5:
            return l_5[c_history.append(current_virtual_move)]
        if c_history[1:].append(current_virtual_move) in l_4:
            return l_4[c_history[1:].append(current_virtual_move)]
        if c_history[2:].append(current_virtual_move) in l_3:
            return l_3[c_history[2:].append(current_virtual_move)]
        if c_history[3:].append(current_virtual_move) in l_4:
            return l_4[c_history[3:].append(current_virtual_move)]
        if c_history[4:].append(current_virtual_move) in l_5:
            return l_5[c_history[4:].append(current_virtual_move)]
        return 0

class local_history:
    record = {}
    def predict(self, row):
        if row[0] not in record:
            return 0
        if row[1] not in record[row[0]]


if __name__=="__main__":
    wb = load_workbook(filename = 'sampletrace.xlsx')
    sheet_ranges = wb['Sheet1']
    predict_threshold = 0
    current_history = ["hitsuccess", "hitsuccess", "hitsuccess", "hitsuccess", "hitsuccess"]


#    print(sheet_ranges['B4'].value)
    for row in sheet_ranges.iter_rows():

        print (row[0].value)
